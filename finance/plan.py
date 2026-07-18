"""Plan ingestion: melt the annual '2026 Cash Flow' sheet (month-across-columns) into
normalised plan_lines. The sheet's month labels are already in our financial-month
naming, so 'Feb' -> period '2026-02' with no offset. Plan lines keep their native
labels; categories.plan_category rolls transaction categories up to them for variance.
"""
from __future__ import annotations

import re
import sqlite3
from pathlib import Path

import openpyxl

from .config import Config
from .normalise import extract_iban

_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], 1)}
_SECTIONS = {"INCOME", "FIXED EXPENSES", "VARIABLE EXPENSES"}
_SKIP_PREFIXES = ("TOTAL", "NET", "BANK", "KEY MILESTONE")
_SKIP_EXACT = {"ILA", "KHALEEJI"}
_PLAN_SHEET_HINT = "cash flow"


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def find_plan_sheet(wb) -> str | None:
    for name in wb.sheetnames:
        if _PLAN_SHEET_HINT in name.lower():
            return name
    return None


def load_plan(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = find_plan_sheet(wb)
    if not sheet:
        return []
    ws = wb[sheet]
    grid = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]

    year = _year_from(grid) or 2026
    header_idx, month_cols, mech_col = _locate_header(grid)
    if header_idx is None:
        return []

    lines: list[dict] = []
    section = None
    for row in grid[header_idx + 1:]:
        c0 = (str(row[0]).strip() if row[0] is not None else "")
        if not c0:
            continue
        up = c0.upper()
        if up in _SECTIONS:
            section = c0
            continue
        if up.startswith(_SKIP_PREFIXES) or up in _SKIP_EXACT:
            continue
        mech = row[mech_col] if mech_col is not None and mech_col < len(row) else None
        mech = str(mech).strip() if mech is not None else None
        for col, mnum in month_cols.items():
            v = row[col] if col < len(row) else None
            if _is_number(v):
                lines.append({
                    "period_id": f"{year}-{mnum:02d}",
                    "category": c0,
                    "section": section,
                    "planned_amount": round(float(v), 3),
                    "mechanism": mech,
                    "mechanism_iban": extract_iban(mech) if mech else None,
                    "source_workbook": path.name,
                })
    return lines


def _year_from(grid) -> int | None:
    for row in grid[:3]:
        for cell in row:
            if isinstance(cell, str):
                m = re.search(r"20\d{2}", cell)
                if m:
                    return int(m.group(0))
    return None


def _locate_header(grid):
    for idx, row in enumerate(grid):
        labels = {str(c).strip().lower(): j for j, c in enumerate(row) if isinstance(c, str)}
        if "jan" in labels and "dec" in labels:
            month_cols = {j: _MONTHS[lab] for lab, j in labels.items() if lab in _MONTHS}
            mech_col = labels.get("mechanism")
            return idx, month_cols, mech_col
    return None, {}, None


def ingest_plan(conn: sqlite3.Connection, cfg: Config, path: Path) -> int:
    from . import db as dbm
    from .periods import period_row

    lines = load_plan(path)
    for ln in lines:
        dbm.ensure_period(conn, period_row(ln["period_id"]))
        conn.execute(
            """INSERT INTO plan_lines(period_id, category, section, planned_amount,
                                      mechanism, mechanism_iban, source_workbook)
               VALUES(:period_id,:category,:section,:planned_amount,
                      :mechanism,:mechanism_iban,:source_workbook)
               ON CONFLICT(period_id, category) DO UPDATE SET
                 planned_amount=excluded.planned_amount, mechanism=excluded.mechanism,
                 mechanism_iban=excluded.mechanism_iban, section=excluded.section,
                 source_workbook=excluded.source_workbook""",
            ln)
    conn.commit()
    return len(lines)
